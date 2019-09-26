#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jun  3 16:28:52 2019

@author: Matteo D'Andrea (s180192)
"""
###################### LYBRARIES AND PACKAGES #################################
import openpyxl
import numpy as np
import pandas as pd
###############################################################################

# a function to extract selected tables from each sheet of a whole excel file 
# input: filename (str) 
# output: pandas dataframe 


def tableImport(filename): 
    
    # the excel file is open 
    # with data_only=True the equations in the cell are ignored
    wb = openpyxl.load_workbook(filename, data_only=True)
    
    ################ SHEETS FILTERING AND TILDE LOCATION ####################
    
    # z is necessary to set a unique id when a sheet has more than one table
    z=0
    # an empty list is initialized 
    key_sheet=[]
    # each cell that starts with ~ in the whole excel file is appended to the list
    # along with the  cell value and coordinates
    for j in wb.sheetnames:
        if j != 'LOG' or j != 'INTRO':
            for row in wb[j].iter_rows():
                for cell in row:
                    if type(cell.value) is str:
                        if cell.value.startswith('~'):
                          cell_row=cell.row
                          cell_col=cell.column
                          key_sheet.append((z,j,cell.value,cell_row,cell_col))
                          z+=1
                            
    # the dictionary will contain each table exported
    dataframe_collection={} 
    
    # for each sheet the same operation are repeated 
    for i in key_sheet:
        
        # the code is a used as unique dictionary key for each table 
        
        code=str(i[0])+'-'+filename.split('.')[0]+'-'+i[1]+i[2]
        sheetname=i[1]
        
        # a specific sheet is selected 
        ws = wb[sheetname]
    
    ##################### TABLE DIMENSIONS AND IMPORT ######################
      
        # from the tilde location the first empty cell on the right is found
        for k in ws.iter_cols(min_col=i[4],min_row=i[3]+1,\
                              max_row=i[3]+1):
                   if k[0].value == None:
                       rightEnd=[(k[0].row,k[0].column-1)]
                       break
                   rightEnd=[(k[0].row,k[0].column)]
                   
        # from the tilde location the first empty cell on the left is found
        for z in reversed(list(ws.iter_cols(max_col=i[4],min_row=i[3]+1,\
                              max_row=i[3]+1))):
                   if z[0].value == None:
                       leftEnd=[(z[0].row,z[0].column+1)]
                       break
                   leftEnd=[(z[0].row,z[0].column)]
    
        # an empty list collects the cells 
        sel=[]
        # given the table width, the rows are collected until an empty row is found
        for s in ws.iter_rows(min_col=leftEnd[0][1],min_row=i[3]+1,\
                                      max_col=rightEnd[0][1]):
            if all(j.value == None for j in s) == True:
                lowerEnd=s[0].row-1
                break 
            else:
                for x in s:
                    sel.append(x.value)
            lowerEnd=s[0].row
        
        # the dataframe is created reshaping the list into a matrix
        matrix=pd.DataFrame(np.reshape(sel,(lowerEnd-i[3],rightEnd[0][1] \
                                            -(leftEnd[0][1]-1))))
        matrix = matrix.applymap(str)
      
        #####################  TABLE SLICING ################################
        # the row and columns which are not necessary are discarded
        selrow=np.char.find(matrix.iloc[:,0].tolist(),'*')==-1
        selcol=np.char.find(matrix.iloc[0,:].tolist(),'*')==-1
        matrix=matrix[selrow]
        matrix=matrix[matrix.columns.values[selcol]]
        selcol2=np.char.find(matrix.iloc[0,:].tolist(),'\I:')==-1
        matrix=matrix[matrix.columns.values[selcol2]]
        # the column names are set 
        matrix.rename(columns=matrix.iloc[0,:],inplace=True)
        matrix=matrix.drop(matrix.index[0])
        # the index is reset after the slicing 
        matrix.reset_index(drop=True,inplace=True)
        # the commodity set column is renamed
        matrix.rename(index=str, columns={'Csets':'Cset','CSet':'Cset'},\
                      inplace=True)
        
        # each table is stored in the dictionary with a unique key that 
        # describes the source and the type of the table
        dataframe_collection[code]=matrix
        
        
    return  dataframe_collection
